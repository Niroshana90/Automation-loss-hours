# adding file save
import openpyxl
import os

source_file = 'BW22 Nov.xlsx'
result_file = 'result.xlsx'

if not os.path.isfile(source_file):
    print('source file dose not exists')
    exit()
else:
    print('source file exists')

wb = openpyxl.load_workbook(source_file)
ws = wb.get_active_sheet()
if os.path.isfile(result_file):
    os.remove(result_file)
    print("removing existing files")

print("creating new result file")
result_wb = openpyxl.Workbook()
written_row_count = int(0)
result_wb_ws = result_wb.get_active_sheet()
result_wb_ws.cell(row=1, column=1).value = ws['A'][3].value
result_wb_ws.cell(row=1, column=2).value = ws['B'][3].value
result_wb_ws.cell(row=1, column=3).value = ws['F'][2].value
result_wb_ws.cell(row=1, column=4).value = ws['G'][2].value
result_wb_ws.cell(row=1, column=5).value = ws['H'][2].value
result_wb_ws.cell(row=1, column=6).value = ws['I'][2].value
result_wb_ws.cell(row=1, column=7).value = ws['J'][2].value
result_wb_ws.cell(row=1, column=8).value = 'Automated mins per GMT'
result_wb_ws.cell(row=1, column=9).value = 'Base mins per GMT'
result_wb_ws.cell(row=1, column=10).value = 'SMV'
result_wb.save(result_file)
written_row_count = 2

for i in range(1, ws.max_row - ws.min_row):
    if ws['B'][i].value == 'Result':

        result_wb_ws.cell(row=written_row_count, column=1).value = ws['A'][i].value
        result_wb_ws.cell(row=written_row_count, column=2).value = ws['B'][i].value
        result_wb_ws.cell(row=written_row_count, column=3).value = ws['F'][i].value
        result_wb_ws.cell(row=written_row_count, column=4).value = ws['G'][i].value
        result_wb_ws.cell(row=written_row_count, column=5).value = ws['H'][i].value
        result_wb_ws.cell(row=written_row_count, column=6).value = ws['I'][i].value
        result_wb_ws.cell(row=written_row_count, column=7).value = ws['J'][i].value
        result_wb_ws.cell(row=written_row_count, column=8).value = result_wb_ws.cell(row=written_row_count, column=5).value/result_wb_ws.cell(row=written_row_count, column=3).value
        result_wb_ws.cell(row=written_row_count, column=9).value = result_wb_ws.cell(row=written_row_count, column=7).value/result_wb_ws.cell(row=written_row_count, column=3).value
        result_wb_ws.cell(row=written_row_count, column=10).value = result_wb_ws.cell(row=written_row_count, column=9).value - result_wb_ws.cell(row=written_row_count, column=8).value

        print('.', sep=' ', end='', flush=True)
        written_row_count += 1

result_wb.save(result_file)